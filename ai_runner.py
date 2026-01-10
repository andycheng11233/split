#!/usr/bin/env python3
"""
Run AI on resolved matches (from auto_matches.json), using current Titan stats and odds.
- Skips matches with no stats or missing data (but NO skip for already-started matches).
- Parallelizes AI calls with configurable concurrency (env AI_CONCURRENCY, default 2, max 10).
- Optional HTTP debug logging (env AI_HTTP_DEBUG=1) to see API status/latency/errors.
- Writes per-match AI JSON, summary, and Excel (bets/exact_scores/not_ready/matches).
- Shows queue progress (start/done counters).
"""

import argparse
import asyncio
import json
import os
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import httpx
import openpyxl
from openpyxl import Workbook

from common_utils import (
    load_titan_stats,
    load_latest_hkjc_odds,
    load_latest_macau_odds,
    has_meaningful_data_for_ai,
    readiness_check,
    extract_time,
    get_kickoff_datetime,
    MACAU_DIR,
    MACAU_DIR_FALLBACK,
    HKJC_DIR,
)

DEEPSEEK_API_KEY = os.getenv("DEEPSEEK_API_KEY")
DEEPSEEK_API_URL = os.getenv("DEEPSEEK_API_URL", "https://api.deepseek.com/chat/completions")
DEEPSEEK_TIMEOUT = float(os.getenv("DEEPSEEK_TIMEOUT", "45"))
DEEPSEEK_RETRIES = int(os.getenv("DEEPSEEK_RETRIES", "3"))
AI_CONCURRENCY = max(1, min(int(os.getenv("AI_CONCURRENCY", "2")), 10))  # cap at 10
AI_HTTP_DEBUG = bool(int(os.getenv("AI_HTTP_DEBUG", "0")))

EXPECTED_TOP_LEVEL_SECTIONS = [
    "match",
    "league_standings",
    "data_comparison_recent10",
    "lineup_and_injuries",
    "last_match_player_ratings",
    "recent10_ratings_parsed",
    "future_matches",
    "head_to_head_sample",
    "league_trend_and_other_stats",
]

# ---------- Helpers ----------


def split_kickoff(kickoff: str) -> Tuple[str, str]:
    if kickoff and "T" in kickoff:
        date_part, time_part = kickoff.split("T", 1)
        return date_part, time_part
    return "", ""


def normalize_parsed_data(parsed: Dict[str, Any]) -> Dict[str, Any]:
    merged = dict(parsed)
    merged["data_comparison_recent10"] = merged.get("data_comparison_recent10") or merged.get("data_comparison")
    merged["recent10_ratings_parsed"] = merged.get("recent10_ratings_parsed") or merged.get("recent_form")
    merged["head_to_head_sample"] = merged.get("head_to_head_sample") or merged.get("head_to_head")

    sections_src = {}
    if isinstance(parsed.get("sections"), dict):
        sections_src.update(parsed["sections"])
    if isinstance(parsed.get("sections_data"), dict):
        sections_src.update(parsed["sections_data"])
    for k, v in sections_src.items():
        merged.setdefault(k, v)

    normalized, missing, available = {}, [], []
    for key in EXPECTED_TOP_LEVEL_SECTIONS:
        val = merged.get(key)
        if val is None:
            normalized[key] = None
            missing.append(key)
        else:
            normalized[key] = val
            available.append(key)
    for k in sections_src.keys():
        if k not in normalized:
            normalized[k] = sections_src[k]
            available.append(k)

    match_block = merged.get("match") or {}
    normalized["match"] = {
        "home_team": match_block.get("home_team") or merged.get("home_team"),
        "away_team": match_block.get("away_team") or merged.get("away_team"),
        "competition": match_block.get("competition") or merged.get("league") or merged.get("competition"),
        "datetime": match_block.get("datetime") or merged.get("game_time") or merged.get("kickoff") or merged.get("datetime"),
        "venue": match_block.get("venue") or merged.get("venue"),
    }

    rr = merged.get("recent10_ratings_parsed") or {}
    normalized["recent10_ratings_parsed"] = {
        "home_recent_ratings_raw": rr.get("home_recent_ratings_raw"),
        "home_recent_ratings": list(rr.get("home_recent_ratings") or []),
        "home_recent_average": rr.get("home_recent_average") or merged.get("home_rating"),
        "away_recent_ratings_raw": rr.get("away_recent_ratings_raw"),
        "away_recent_ratings": list(rr.get("away_recent_ratings") or []),
        "away_recent_average": rr.get("away_recent_average") or merged.get("away_rating"),
    }

    section_counts = {}
    for key, v in normalized.items():
        if key.startswith("_"):
            continue
        if isinstance(v, list):
            section_counts[key] = len(v)
        elif isinstance(v, dict):
            section_counts[key] = len(v.keys())
        elif v is None:
            section_counts[key] = 0
        else:
            section_counts[key] = 1

    normalized["_meta"] = {
        "missing_fields": missing,
        "available_sections": list(dict.fromkeys(available)),
        "section_counts": section_counts,
    }
    return normalized


def truncate_str(x: Any, max_len: int = 240) -> Any:
    if isinstance(x, str) and len(x) > max_len:
        return x[:max_len] + "…"
    return x


def slim_sections_data(sections_data: Dict[str, Any], max_rows: int = 6, max_str_len: int = 240) -> Dict[str, Any]:
    if not isinstance(sections_data, dict):
        return sections_data
    out = {}
    for key, val in sections_data.items():
        if isinstance(val, list):
            rows = val[:max_rows]
            new_rows = []
            for row in rows:
                if isinstance(row, list):
                    new_rows.append([truncate_str(c, max_str_len) for c in row])
                else:
                    new_rows.append(truncate_str(row, max_str_len))
            out[key] = new_rows
        else:
            out[key] = val
    return out


def slim_stats_for_prompt(stats: Dict[str, Any], max_rows: int = 6, max_str_len: int = 240) -> Dict[str, Any]:
    slim = json.loads(json.dumps(stats))
    if "sections_data" in slim and isinstance(slim["sections_data"], dict):
        slim["sections_data"] = slim_sections_data(
            slim["sections_data"], max_rows=max_rows, max_str_len=max_str_len
        )
    return slim


def slim_hkjc_markets(markets: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(markets, dict):
        return markets
    kept = {}
    allow_keys = {
        "home_odds",
        "draw_odds",
        "away_odds",
        "over_odds",
        "under_odds",
        "goal_line",
        "asia_handicap_value",
        "asia_handicap_give_home",
        "euro_handicap_value",
        "euro_handicap_give_home",
        "handicap",
        "line_index",
        "score",
        "odds",
        "scores",
        "buckets",
        "combos",
    }
    for mkt, val in markets.items():
        if isinstance(val, dict):
            if "scores" in val and isinstance(val.get("scores"), list):
                kept[mkt] = {"scores": val["scores"]}
                continue
            if "buckets" in val and isinstance(val.get("buckets"), list):
                kept[mkt] = {"buckets": val["buckets"]}
                continue
            if "combos" in val and isinstance(val.get("combos"), list):
                kept[mkt] = {"combos": val["combos"]}
                continue
            kept[mkt] = {k: v for k, v in val.items() if k in allow_keys}
        elif isinstance(val, list):
            new_list = []
            for item in val:
                if isinstance(item, dict):
                    new_list.append({k: v for k, v in item.items() if k in allow_keys})
                else:
                    new_list.append(item)
            kept[mkt] = new_list
        else:
            kept[mkt] = val
    return kept


def slim_macau_markets(markets: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(markets, dict):
        return markets
    kept = {}
    for mkt, val in markets.items():
        if not isinstance(val, dict):
            kept[mkt] = val
            continue
        slim = {}
        if "market_name" in val:
            slim["market_name"] = val["market_name"]
        if "all_numbers" in val:
            slim["all_numbers"] = val["all_numbers"]
        if "all_lines" in val:
            slim["all_lines"] = val["all_lines"]
        for k in ("scores", "buckets", "combos"):
            if k in val:
                slim[k] = val[k]
        kept[mkt] = slim
    return kept


def slim_odds_bundle(odds_bundle: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not odds_bundle:
        return None
    out = {}
    for src, obj in odds_bundle.items():
        if not isinstance(obj, dict):
            out[src] = obj
            continue
        slim = {}
        for k in (
            "event_id",
            "match",
            "home_team",
            "away_team",
            "tournament",
            "time_raw",
            "kickoff",
            "kickoff_date",
            "kickoff_time",
        ):
            if k in obj:
                slim[k] = obj[k]
        markets = obj.get("markets")
        if markets:
            if src == "hkjc":
                slim["markets"] = slim_hkjc_markets(markets)
            else:
                slim["markets"] = slim_macau_markets(markets)
        out[src] = slim
    return out


def estimate_tokens(prompt: str, chinese_heavy: bool = True) -> int:
    return int(len(prompt) * (0.9 if chinese_heavy else 0.25))


def build_ai_prompt_full(normalized: dict, use_chinese: bool = True, per_section_limit: int = 2000) -> str:
    meta = normalized.get("_meta", {})
    available = meta.get("available_sections", [])
    missing = meta.get("missing_fields", [])
    odds_bundle = normalized.get("_odds_bundle")

    def render(name: str, limit: int = per_section_limit) -> str:
        if name not in normalized:
            return ""
        try:
            s = json.dumps(normalized[name], ensure_ascii=False, indent=0)
        except Exception:
            s = str(normalized[name])
        if name != "_odds_bundle" and len(s) > limit:
            s = s[:limit] + "... (truncated)"
        return f"=== {name} ===\n{s}"

    contract = (
        '{ "bets": [ { "market": "...", "line": "", "selection": "...", "price": 0, "bookmaker": "...", '
        '"confidence": 1-10, "value_flag": "value|fair|pass", "reason": "..." } ], '
        '"exact_scores": [ { "score": "1-0", "confidence": 1-5, "reason": "..." } ] }'
    )

    if use_chinese:
        header = (
            f"你是一位資深足球博彩分析師。可用欄位: {', '.join(available) or 'none'}；缺失欄位: {', '.join(missing) or 'none'}。"
            f" {'Odds bundle present: ' + ', '.join(odds_bundle.keys()) if odds_bundle else ''}"
            " 對每個可用市場，提供最佳投注：如同一市場/盤口/選項在多家（如 hkjc、macau）都有，請只保留賠率最高的，並在 bookmaker 標明來源。"
            " 同一市場可有不同盤口的多個選項，但同盤口/選項只保留賠率最高的一條。按價值排序。"
            " 僅輸出 JSON，格式：\n" + contract
        )
    else:
        header = (
            f"Available: {', '.join(available) or 'none'}; missing: {', '.join(missing) or 'none'}. "
            f"{'Odds bundle: ' + ', '.join(odds_bundle.keys()) if odds_bundle else ''} "
            "For each market, return the best-priced picks: if the same market/line/selection appears at multiple books (hkjc/macau), keep only the highest price and set bookmaker accordingly. "
            "A market may have multiple lines, but for the same line/selection keep only the top-priced one. Order by value. "
            "Output JSON only in this shape:\n" + contract
        )

    parts = []
    if odds_bundle:
        parts.append(render("_odds_bundle", limit=999_999_999))
    for k in sorted(normalized.keys()):
        if k.startswith("_"):
            continue
        section = render(k, per_section_limit)
        if section:
            parts.append(section)

    return header + "\n\n" + "\n\n".join(parts)


def parse_ai_json_response(text: str) -> Optional[Dict[str, Any]]:
    if not text:
        return None
    try:
        return json.loads(text)
    except Exception:
        pass
    m = re.search(r"(\{(?:.|\s)*\})", text)
    if m:
        try:
            return json.loads(m.group(1))
        except Exception:
            return None
    return None


async def call_deepseek_api_async(prompt: str, timeout=None, max_retries=None) -> str:
    if timeout is None:
        timeout = int(DEEPSEEK_TIMEOUT)
    if max_retries is None:
        max_retries = int(DEEPSEEK_RETRIES)

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {DEEPSEEK_API_KEY}" if DEEPSEEK_API_KEY else "",
    }
    payload = {
        "model": "deepseek-chat",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.7,
    }

    backoff_base = 0.6
    last_err = None
    async with httpx.AsyncClient(timeout=timeout) as client:
        for attempt in range(1, max_retries + 1):
            t0 = time.time()
            if AI_HTTP_DEBUG:
                print(f"[AI][http] attempt {attempt}/{max_retries} len_prompt={len(prompt)}")
            try:
                resp = await client.post(DEEPSEEK_API_URL, headers=headers, json=payload)
                elapsed = time.time() - t0
                if AI_HTTP_DEBUG:
                    print(f"[AI][http] status={resp.status_code} elapsed={elapsed:.2f}s content_len={len(resp.text)}")
                resp.raise_for_status()
                data = resp.json()
                if isinstance(data, dict):
                    choices = data.get("choices")
                    if choices and isinstance(choices, list) and len(choices) > 0:
                        first = choices[0]
                        if isinstance(first, dict):
                            msg = first.get("message") or first.get("text") or {}
                            if isinstance(msg, dict):
                                content = msg.get("content") or msg.get("text") or ""
                            elif isinstance(msg, str):
                                content = msg
                            else:
                                content = ""
                        else:
                            content = str(first)
                        return content.strip()
                return resp.text
            except httpx.HTTPStatusError as e:
                last_err = str(e)
                status = getattr(e.response, "status_code", None)
                if AI_HTTP_DEBUG:
                    print(f"[AI][http] HTTP error status={status} err={repr(e)}")
                if status and status < 500 and status != 429:
                    break
            except Exception as e:
                last_err = str(e)
                if AI_HTTP_DEBUG:
                    print(f"[AI][http] exception err={repr(e)}")
            await asyncio.sleep(backoff_base * attempt)
    raise RuntimeError(f"DeepSeek API calls failed: {last_err}")


async def perform_ai_analysis_for_match_async(normalized: dict, use_chinese: bool = True) -> dict:
    result = {
        "bets": [],
        "exact_scores": [],
        "ai_raw_response": None,
        "ai_parsed_json": None,
        "data_availability": normalized.get("_meta", {}),
        "prompt": None,
    }

    if not has_meaningful_data_for_ai(normalized):
        result["ai_raw_response"] = "No meaningful stats; skipped AI call."
        return result

    prompt = build_ai_prompt_full(normalized, use_chinese=use_chinese, per_section_limit=2000)
    result["prompt"] = prompt
    print(f"[AI] Prompt chars: {len(prompt)} | est tokens: {estimate_tokens(prompt, chinese_heavy=use_chinese)}")

    last_err = None
    for attempt in range(1, 3):
        try:
            ai_text = await call_deepseek_api_async(prompt)
            result["ai_raw_response"] = ai_text
            parsed = parse_ai_json_response(ai_text)
            if parsed and ("bets" in parsed or "exact_scores" in parsed):
                result["ai_parsed_json"] = parsed
                result["bets"] = parsed.get("bets", [])
                result["exact_scores"] = parsed.get("exact_scores", [])
                return result
            last_err = "No parsable bets/exact_scores in response"
        except Exception as e:
            last_err = str(e)
        await asyncio.sleep(0.6 * attempt)

    result["ai_raw_response"] = f"AI call failed: {last_err}"
    return result


# ---------- Excel writer ----------


def write_excel_report(
    path: Path,
    bets_rows: List[List[Any]],
    score_rows: List[List[Any]],
    not_ready_rows: List[List[Any]],
    matches_rows: List[List[Any]],
):
    wb = Workbook()
    ws_bets = wb.active
    ws_bets.title = "bets"
    ws_bets.append(
        [
            "titan_id",
            "home",
            "away",
            "kickoff_date",
            "kickoff_time",
            "sources",
            "market",
            "line",
            "selection",
            "price",
            "bookmaker",
            "confidence",
            "value_flag",
            "reason",
            "ai_status",
        ]
    )
    for r in bets_rows:
        ws_bets.append(r)

    ws_scores = wb.create_sheet("exact_scores")
    ws_scores.append(
        ["titan_id", "home", "away", "kickoff_date", "kickoff_time", "sources", "score", "confidence", "reason", "ai_status"]
    )
    for r in score_rows:
        ws_scores.append(r)

    ws_nr = wb.create_sheet("not_ready")
    ws_nr.append(["titan_id", "home", "away", "kickoff_date", "kickoff_time", "reasons"])
    for r in not_ready_rows:
        ws_nr.append(r)

    ws_matches = wb.create_sheet("matches")
    ws_matches.append(
        [
            "titan_id",
            "home",
            "away",
            "kickoff_date",
            "kickoff_time",
            "sources",
            "ai_status",
            "bets_count",
            "scores_count",
            "reasons",
        ]
    )
    for r in matches_rows:
        ws_matches.append(r)

    wb.save(path)


# ---------- Per-match pipeline ----------


async def process_match(
    m: Dict[str, Any],
    titan_map: Dict[str, Any],
    hkjc_map: Dict[str, Any],
    macau_map: Dict[str, Any],
    api_key_present: bool,
    out_dir: Path,
    force: bool,
    sem: asyncio.Semaphore,
    progress: Dict[str, int],
    progress_lock: asyncio.Lock,
    total: int,
) -> Tuple[
    List[Dict[str, Any]],
    List[List[Any]],
    List[List[Any]],
    List[List[Any]],
    List[List[Any]],
]:
    """
    Process a single match; returns (summary, bets_rows, score_rows, not_ready_rows, matches_rows).
    """
    summary: List[Dict[str, Any]] = []
    bets_rows: List[List[Any]] = []
    score_rows: List[List[Any]] = []
    not_ready_rows: List[List[Any]] = []
    matches_rows: List[List[Any]] = []

    async with sem:
        async with progress_lock:
            progress["started"] += 1
            print(f"[AI][start] {progress['started']}/{total} titan_id={m.get('titan_id')}")
        t_start = time.time()
        try:
            tid = m["titan_id"]
            stats = titan_map.get(tid)
            home = ""
            away = ""
            kickoff = ""
            sources = ""

            if stats:
                home = stats.get("home_team") or stats.get("match", {}).get("home_team") or ""
                away = stats.get("away_team") or stats.get("match", {}).get("away_team") or ""
                kickoff_dt = get_kickoff_datetime(stats)
                kickoff = kickoff_dt.isoformat() if kickoff_dt else extract_time(stats) or ""
            else:
                kickoff_dt = None

            kickoff_date, kickoff_time = split_kickoff(kickoff)

            # No stats at all -> precheck fail
            if not stats:
                reasons = ["no_stats_file"]
                not_ready_rows.append([tid, home, away, kickoff_date, kickoff_time, ";".join(reasons)])
                summary.append({"titan_id": tid, "status": "precheck_fail", "reasons": reasons})
                matches_rows.append([tid, home, away, kickoff_date, kickoff_time, sources, "precheck_fail", 0, 0, ";".join(reasons)])
                return summary, bets_rows, score_rows, not_ready_rows, matches_rows

            stats_for_prompt = slim_stats_for_prompt(stats)
            normalized = normalize_parsed_data(stats_for_prompt)

            odds_bundle = {}
            hid = m.get("hkjc_event_id")
            mid = m.get("macau_event_id")
            if hid and hid in hkjc_map:
                odds_bundle["hkjc"] = hkjc_map[hid]
            if mid and mid in macau_map:
                odds_bundle["macau"] = macau_map[mid]

            slimmed_odds = slim_odds_bundle(odds_bundle) if odds_bundle else None
            normalized["_odds_bundle"] = slimmed_odds if slimmed_odds else None
            sources = ",".join(sorted(slimmed_odds.keys())) if slimmed_odds else ""

            ready, reasons = readiness_check(normalized, slimmed_odds, api_key_present)

            if not ready:
                not_ready_rows.append([tid, home, away, kickoff_date, kickoff_time, ";".join(reasons)])
                summary.append(
                    {
                        "titan_id": tid,
                        "hkjc_event_id": hid,
                        "macau_event_id": mid,
                        "status": "precheck_fail",
                        "reasons": reasons,
                    }
                )
                matches_rows.append([tid, home, away, kickoff_date, kickoff_time, sources, "precheck_fail", 0, 0, ";".join(reasons)])
                return summary, bets_rows, score_rows, not_ready_rows, matches_rows

            out_path = out_dir / f"ai_{tid}.json"
            if out_path.exists() and not force:
                print(f"[AI][cache] {tid} using {out_path}")
                try:
                    cached = json.loads(out_path.read_text(encoding="utf-8"))
                except Exception:
                    cached = {}
                cached_bets = cached.get("bets") or []
                cached_scores = cached.get("exact_scores") or []
                ai_status = "cached"
                summary.append(
                    {
                        "titan_id": tid,
                        "hkjc_event_id": hid,
                        "macau_event_id": mid,
                        "status": ai_status,
                        "bets": len(cached_bets),
                    }
                )
                for b in cached_bets:
                    bets_rows.append(
                        [
                            tid,
                            home,
                            away,
                            kickoff_date,
                            kickoff_time,
                            sources,
                            b.get("market", ""),
                            b.get("line", ""),
                            b.get("selection", ""),
                            b.get("price", ""),
                            b.get("bookmaker", ""),
                            b.get("confidence", ""),
                            b.get("value_flag", ""),
                            b.get("reason", ""),
                            ai_status,
                        ]
                    )
                for s in cached_scores:
                    score_rows.append(
                        [
                            tid,
                            home,
                            away,
                            kickoff_date,
                            kickoff_time,
                            sources,
                            s.get("score", ""),
                            s.get("confidence", ""),
                            s.get("reason", ""),
                            ai_status,
                        ]
                    )
                matches_rows.append([tid, home, away, kickoff_date, kickoff_time, sources, ai_status, len(cached_bets), len(cached_scores), ""])
                summary[-1]["cached_bets"] = cached_bets
                summary[-1]["cached_exact_scores"] = cached_scores
                return summary, bets_rows, score_rows, not_ready_rows, matches_rows

            # Run AI
            res = await perform_ai_analysis_for_match_async(normalized, use_chinese=True)
            print(f"[AI][write] {tid} -> {out_path}")
            out_path.write_text(json.dumps(res, ensure_ascii=False, indent=2), encoding="utf-8")

            ai_status = "ok" if res.get("bets") or res.get("exact_scores") else "no_parsed_ai"
            summary.append(
                {
                    "titan_id": tid,
                    "hkjc_event_id": hid,
                    "macau_event_id": mid,
                    "status": ai_status,
                    "bets": len(res.get("bets", [])),
                }
            )

            for b in res.get("bets", []):
                bets_rows.append(
                    [
                        tid,
                        home,
                        away,
                        kickoff_date,
                        kickoff_time,
                        sources,
                        b.get("market", ""),
                        b.get("line", ""),
                        b.get("selection", ""),
                        b.get("price", ""),
                        b.get("bookmaker", ""),
                        b.get("confidence", ""),
                        b.get("value_flag", ""),
                        b.get("reason", ""),
                        ai_status,
                    ]
                )

            for s in res.get("exact_scores", []):
                score_rows.append(
                    [
                        tid,
                        home,
                        away,
                        kickoff_date,
                        kickoff_time,
                        sources,
                        s.get("score", ""),
                        s.get("confidence", ""),
                        s.get("reason", ""),
                        ai_status,
                    ]
                )

            matches_rows.append(
                [
                    tid,
                    home,
                    away,
                    kickoff_date,
                    kickoff_time,
                    sources,
                    ai_status,
                    len(res.get("bets", [])),
                    len(res.get("exact_scores", [])),
                    "" if ai_status == "ok" else "no_parsed_ai",
                ]
            )
        finally:
            elapsed = time.time() - t_start
            async with progress_lock:
                progress["done"] += 1
                print(f"[AI][done ] {progress['done']}/{total} titan_id={m.get('titan_id')} in {elapsed:.1f}s")
    return summary, bets_rows, score_rows, not_ready_rows, matches_rows


# ---------- Main ----------


async def main_async(out_dir: Path, excel_path: Path, force: bool, only_ids: Optional[str], redo_missing: bool):
    total_start = time.time()
    out_dir.mkdir(parents=True, exist_ok=True)
    matches_path = out_dir / "auto_matches.json"
    if not matches_path.exists():
        raise FileNotFoundError(f"{matches_path} not found. Run match_resolve.py first.")

    matches = json.loads(matches_path.read_text(encoding="utf-8"))

    # Filter selection
    only_set = set(x.strip() for x in only_ids.split(",")) if only_ids else None
    selected: List[Dict[str, Any]] = []
    for m in matches:
        tid = str(m.get("titan_id") or "").strip()
        if not tid:
            continue
        if only_set and tid not in only_set:
            continue
        if redo_missing and not force:
            out_path = out_dir / f"ai_{tid}.json"
            if out_path.exists():
                try:
                    cached = json.loads(out_path.read_text(encoding="utf-8"))
                    if (cached.get("bets") or cached.get("exact_scores")):
                        continue  # already populated; skip
                except Exception:
                    pass
        selected.append(m)

    if not selected:
        print("No matches selected after filters; exiting.")
        return

    titan_map = load_titan_stats()
    hkjc_map = load_latest_hkjc_odds(HKJC_DIR)
    macau_map = load_latest_macau_odds(MACAU_DIR, MACAU_DIR_FALLBACK)

    api_key_present = bool(DEEPSEEK_API_KEY)
    sem = asyncio.Semaphore(AI_CONCURRENCY)
    progress = {"started": 0, "done": 0}
    progress_lock = asyncio.Lock()
    total = len(selected)

    tasks = [
        process_match(
            m,
            titan_map,
            hkjc_map,
            macau_map,
            api_key_present,
            out_dir,
            force,
            sem,
            progress,
            progress_lock,
            total,
        )
        for m in selected
    ]
    results = await asyncio.gather(*tasks)

    summary: List[Dict[str, Any]] = []
    bets_rows: List[List[Any]] = []
    score_rows: List[List[Any]] = []
    not_ready_rows: List[List[Any]] = []
    matches_rows: List[List[Any]] = []

    for s, b, sc, nr, mr in results:
        summary.extend(s)
        bets_rows.extend(b)
        score_rows.extend(sc)
        not_ready_rows.extend(nr)
        matches_rows.extend(mr)

    # Sort matches by kickoff descending (latest first)
    def _dt_key(row: List[Any]) -> str:
        # row indexes: 3=date, 4=time
        return (row[3] or "") + "T" + (row[4] or "")

    matches_rows.sort(key=_dt_key, reverse=True)

    write_excel_report(excel_path, bets_rows, score_rows, not_ready_rows, matches_rows)

    precheck_fails = sum(1 for s in summary if s.get("status") == "precheck_fail")
    ready_count = total - precheck_fails
    summary_path = out_dir / f"ai_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    total_elapsed = time.time() - total_start
    print(f"Done. Matches: {total} | Ready (passed gate): {ready_count} | Not ready: {precheck_fails}")
    print(f"Summary: {summary_path} | Excel: {excel_path}")
    print(f"Total elapsed: {total_elapsed:.1f}s")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="out_ai_auto", help="Output directory (should contain auto_matches.json)")
    ap.add_argument("--excel", default="ai_report.xlsx", help="Excel report path")
    ap.add_argument("--force", action="store_true", help="Reprocess even if ai_*.json exists")
    ap.add_argument("--only-ids", help="Comma-separated titan_ids to process; skip others")
    ap.add_argument(
        "--redo-missing",
        action="store_true",
        help="Only rerun matches whose ai_*.json is missing or has no bets/scores",
    )
    args = ap.parse_args()
    asyncio.run(
        main_async(
            Path(args.out),
            Path(args.excel),
            args.force,
            args.only_ids,
            args.redo_missing,
        )
    )


if __name__ == "__main__":
    main()
